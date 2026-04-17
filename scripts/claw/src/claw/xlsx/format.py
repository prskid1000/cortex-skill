"""claw xlsx format — set Excel number format on a cell range."""

from __future__ import annotations

import re
from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, RangeSelector, common_output_options, die, emit_json, safe_write,
)


@click.command(name="format")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True)
@click.option("--range", "a1_range", required=True,
              help='A1 range; also accepts whole columns like "B:B" or "B:D".')
@click.option("--number-format", "number_format", required=True,
              help='e.g. "#,##0.00", "0%", "$#,##0.00", "yyyy-mm-dd".')
@common_output_options
def format(src: Path, sheet: str, a1_range: str, number_format: str,
           force: bool, backup: bool, as_json: bool, dry_run: bool,
           quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Set `cell.number_format` on every cell in --range."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    if dry_run:
        click.echo(f"would set number_format={number_format!r} on {sheet}!{a1_range}")
        return

    wb = load_workbook(src)
    ws = wb[sheet]

    col_m = re.match(r"^([A-Za-z]+)(?::([A-Za-z]+))?$", a1_range.strip())
    touched = 0
    if col_m:
        c1 = column_index_from_string(col_m.group(1).upper())
        c2 = column_index_from_string((col_m.group(2) or col_m.group(1)).upper())
        last_row = ws.max_row
        for col_idx in range(c1, c2 + 1):
            for row_idx in range(1, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = number_format
                touched += 1
    else:
        r1, c1, r2, c2 = RangeSelector(a1_range).resolve()
        r2 = r2 or r1
        c2 = c2 or c1
        for row_idx in range(r1, r2 + 1):
            for col_idx in range(c1, c2 + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = number_format
                touched += 1

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "sheet": sheet, "range": a1_range,
                   "number_format": number_format, "cells": touched})
    elif not quiet:
        click.echo(f"formatted {touched} cells on {sheet}!{a1_range}")
