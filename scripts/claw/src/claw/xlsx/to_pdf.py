"""claw xlsx to-pdf — render a workbook to PDF via LibreOffice headless."""

from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, EXIT_SYSTEM, common_output_options, die, emit_json, safe_copy,
)


@click.command(name="to-pdf")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--out", "out_path", required=True, type=click.Path(path_type=Path))
@click.option("--sheet", "sheets", multiple=True,
              help="Sheet name (repeat for more). Unused unless using --fit-to-width.")
@click.option("--fit-to-width", "fit_to_width", is_flag=True,
              help="Apply fit-to-1-page-wide on each sheet before converting.")
@click.option("--orientation", default=None,
              type=click.Choice(["portrait", "landscape"]))
@common_output_options
def to_pdf(src: Path, out_path: Path, sheets: tuple[str, ...],
           fit_to_width: bool, orientation: str | None,
           force: bool, backup: bool, as_json: bool, dry_run: bool,
           quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Render xlsx to PDF via LibreOffice (soffice) headless."""
    soffice = _find_soffice()
    if not soffice:
        die("libreoffice / soffice not found on PATH",
            code=EXIT_INPUT,
            hint="install LibreOffice (https://www.libreoffice.org/) and "
                 "make sure `soffice` is on PATH",
            as_json=as_json)

    if dry_run:
        click.echo(f"would render {src} -> {out_path} via {soffice}")
        return

    working_src = src
    if fit_to_width or orientation:
        working_src = _pre_adjust(src, sheets, fit_to_width, orientation, as_json=as_json)

    with tempfile.TemporaryDirectory(prefix="claw-topdf-") as td:
        td_path = Path(td)
        try:
            proc = subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf",
                 "--outdir", str(td_path), str(working_src)],
                capture_output=True, text=True, timeout=180,
            )
        except FileNotFoundError:
            die(f"could not execute {soffice}", code=EXIT_SYSTEM, as_json=as_json)
        if proc.returncode != 0:
            die(f"soffice failed: {proc.stderr or proc.stdout}",
                code=EXIT_SYSTEM, as_json=as_json)

        produced = next(iter(td_path.glob("*.pdf")), None)
        if produced is None:
            die("soffice reported success but no PDF was produced",
                code=EXIT_SYSTEM, as_json=as_json)

        if out_path.exists() and not force:
            die(f"{out_path} exists (pass --force)",
                code=EXIT_INPUT, as_json=as_json)
        safe_copy(produced, out_path, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(out_path), "src": str(src)})
    elif not quiet:
        click.echo(f"wrote {out_path}")


def _find_soffice() -> str | None:
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    for cand in (
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ):
        if Path(cand).exists():
            return cand
    return None


def _pre_adjust(src: Path, sheets: tuple[str, ...], fit_to_width: bool,
                orientation: str | None, *, as_json: bool) -> Path:
    """Clone src to a temp path with print-setup tweaks applied."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    tmp_dir = Path(tempfile.mkdtemp(prefix="claw-topdf-"))
    tmp_src = tmp_dir / src.name
    shutil.copy2(src, tmp_src)

    wb = load_workbook(tmp_src)
    target_sheets = list(sheets) if sheets else wb.sheetnames
    for name in target_sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        if fit_to_width:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        if orientation:
            ws.page_setup.orientation = orientation
    wb.save(tmp_src)
    return tmp_src
