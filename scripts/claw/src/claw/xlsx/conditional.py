"""claw xlsx conditional — add conditional-formatting rules."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import EXIT_INPUT, common_output_options, die, emit_json, safe_write


@click.command(name="conditional")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True)
@click.option("--range", "a1_range", required=True)
@click.option("--cell-is", "cell_is", default=None,
              help='Op:val e.g. "greaterThan:100" or "between:0:10".')
@click.option("--formula", default=None, help="Custom =FORMULA() expression.")
@click.option("--color-scale", "color_scale", default=None,
              help='Shorthand like "min:#F8696B,max:#63BE7B".')
@click.option("--data-bar", "data_bar", default=None, help="Single hex, e.g. #638EC6.")
@click.option("--icon-set", "icon_set", default=None,
              help='"3TrafficLights1:percent:0,33,67"')
@click.option("--fill", default=None, help="#RRGGBB for style fill.")
@click.option("--color", default=None, help="#RRGGBB font color.")
@click.option("--bold", is_flag=True)
@click.option("--stop-if-true", "stop_if_true", is_flag=True)
@common_output_options
def conditional(src: Path, sheet: str, a1_range: str,
                cell_is: str | None, formula: str | None,
                color_scale: str | None, data_bar: str | None, icon_set: str | None,
                fill: str | None, color: str | None, bold: bool, stop_if_true: bool,
                force: bool, backup: bool, as_json: bool, dry_run: bool,
                quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Add one conditional-formatting rule (pick exactly one rule kind)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import (
            CellIsRule, ColorScaleRule, DataBarRule, FormulaRule, IconSetRule,
        )
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    picked = [x for x in (cell_is, formula, color_scale, data_bar, icon_set) if x]
    if len(picked) != 1:
        die("specify exactly one rule kind", code=EXIT_INPUT, as_json=as_json)

    if dry_run:
        click.echo(f"would add rule on {sheet}!{a1_range}")
        return

    wb = load_workbook(src)
    ws = wb[sheet]

    dxf_fill = PatternFill(start_color=fill.lstrip("#"), end_color=fill.lstrip("#"),
                           fill_type="solid") if fill else None
    dxf_font = Font(bold=bold or None, color=color.lstrip("#")) if (color or bold) else None

    rule = None
    if cell_is:
        parts = cell_is.split(":")
        op = parts[0]
        formulas = parts[1:]
        rule = CellIsRule(operator=op, formula=formulas, stopIfTrue=stop_if_true,
                          fill=dxf_fill, font=dxf_font)
    elif formula:
        rule = FormulaRule(formula=[formula.lstrip("=")], stopIfTrue=stop_if_true,
                           fill=dxf_fill, font=dxf_font)
    elif color_scale:
        parts = dict(p.split(":") for p in color_scale.split(","))
        rule = ColorScaleRule(
            start_type="min", start_color=parts.get("min", "#F8696B").lstrip("#"),
            end_type="max", end_color=parts.get("max", "#63BE7B").lstrip("#"),
        )
    elif data_bar:
        rule = DataBarRule(start_type="min", end_type="max",
                           color=data_bar.lstrip("#"), showValue=True)
    elif icon_set:
        parts = icon_set.split(":")
        kind = parts[0]
        scope = "percent"
        thresholds: list[float] = [0, 33, 67]
        if len(parts) >= 3 and parts[1] in ("percent", "num", "percentile", "formula"):
            scope = parts[1]
            try:
                thresholds = [float(v) for v in parts[2].split(",") if v.strip()]
            except ValueError:
                die(f"invalid icon-set thresholds: {parts[2]!r}",
                    code=EXIT_INPUT, as_json=as_json)
        elif len(parts) == 2 and parts[1]:
            try:
                thresholds = [float(v) for v in parts[1].split(",") if v.strip()]
            except ValueError:
                die(f"invalid icon-set thresholds: {parts[1]!r}",
                    code=EXIT_INPUT, as_json=as_json)
        rule = IconSetRule(kind, scope, thresholds)

    ws.conditional_formatting.add(a1_range, rule)

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "sheet": sheet, "range": a1_range})
    elif not quiet:
        click.echo(f"added conditional rule on {sheet}!{a1_range}")
