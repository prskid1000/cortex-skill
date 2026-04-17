"""claw xlsx style — apply font / fill / border / alignment to a range."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, RangeSelector, common_output_options, die, emit_json, safe_write,
)


PRESETS = {
    "header": {"bold": True, "color": "#FFFFFF", "fill": "#4472C4",
               "halign": "center", "border": "thin"},
    "total":  {"bold": True, "fill": "#D9E1F2", "border": "thin"},
    "alt-rows": {"fill": "#F2F2F2"},
    "zebra":  {"fill": "#F2F2F2"},
}


@click.command(name="style")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True)
@click.option("--range", "a1_range", required=True)
@click.option("--preset", default=None,
              type=click.Choice(list(PRESETS.keys())),
              help="Canned style — individual flags override preset fields.")
@click.option("--bold", is_flag=True)
@click.option("--italic", is_flag=True)
@click.option("--size", "font_size", default=None, type=int)
@click.option("--font", "font_name", default=None)
@click.option("--color", default=None, help="#RRGGBB font color.")
@click.option("--fill", default=None, help="#RRGGBB cell fill.")
@click.option("--border", default=None,
              type=click.Choice(["thin", "thick", "double", "medium", "dotted", "dashed"]))
@click.option("--border-color", "border_color", default=None)
@click.option("--halign", default=None,
              type=click.Choice(["left", "center", "right"]))
@click.option("--valign", default=None,
              type=click.Choice(["top", "center", "bottom"]))
@click.option("--wrap", is_flag=True)
@common_output_options
def style(src: Path, sheet: str, a1_range: str, preset: str | None,
          bold: bool, italic: bool, font_size: int | None, font_name: str | None,
          color: str | None, fill: str | None, border: str | None,
          border_color: str | None, halign: str | None, valign: str | None,
          wrap: bool,
          force: bool, backup: bool, as_json: bool, dry_run: bool,
          quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Apply font / fill / border / alignment to every cell in --range."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    opts = {"bold": bold or None, "italic": italic or None,
            "size": font_size, "font": font_name,
            "color": color, "fill": fill, "border": border,
            "border_color": border_color, "halign": halign, "valign": valign,
            "wrap": wrap or None}
    if preset:
        for k, v in PRESETS[preset].items():
            if opts.get(k) in (None, False):
                opts[k] = v

    if dry_run:
        click.echo(f"would apply style to {sheet}!{a1_range}: {opts}")
        return

    wb = load_workbook(src)
    ws = wb[sheet]
    r1, c1, r2, c2 = RangeSelector(a1_range).resolve()
    r2 = r2 or r1
    c2 = c2 or c1

    font_kwargs = {}
    if opts["bold"]:
        font_kwargs["bold"] = True
    if opts["italic"]:
        font_kwargs["italic"] = True
    if opts["size"]:
        font_kwargs["size"] = opts["size"]
    if opts["font"]:
        font_kwargs["name"] = opts["font"]
    if opts["color"]:
        font_kwargs["color"] = opts["color"].lstrip("#")

    pat_fill = None
    if opts["fill"]:
        hx = opts["fill"].lstrip("#")
        pat_fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")

    brd = None
    if opts["border"]:
        side = Side(style=opts["border"],
                    color=(opts["border_color"] or "#000000").lstrip("#"))
        brd = Border(left=side, right=side, top=side, bottom=side)

    align = None
    if opts["halign"] or opts["valign"] or opts["wrap"]:
        align = Alignment(
            horizontal=opts["halign"],
            vertical=("center" if opts["valign"] == "center" else opts["valign"]),
            wrap_text=bool(opts["wrap"]),
        )

    alt = preset in ("alt-rows", "zebra")

    for row_idx in range(r1, r2 + 1):
        if alt and (row_idx - r1) % 2 == 0:
            continue
        for col_idx in range(c1, c2 + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if font_kwargs:
                existing = cell.font
                cell.font = Font(
                    name=font_kwargs.get("name", existing.name),
                    size=font_kwargs.get("size", existing.size),
                    bold=font_kwargs.get("bold", existing.bold),
                    italic=font_kwargs.get("italic", existing.italic),
                    color=font_kwargs.get("color", existing.color),
                )
            if pat_fill:
                cell.fill = pat_fill
            if brd:
                cell.border = brd
            if align:
                cell.alignment = align

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "sheet": sheet, "range": a1_range,
                   "preset": preset})
    elif not quiet:
        click.echo(f"styled {sheet}!{a1_range}" + (f" (preset={preset})" if preset else ""))
