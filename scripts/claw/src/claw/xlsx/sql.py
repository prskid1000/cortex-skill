"""claw xlsx sql — run SQL against sheets (DuckDB preferred, SQLite fallback)."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, common_output_options, die, emit_json, safe_write, write_rows_csv,
)


@click.command(name="sql")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.argument("query")
@click.option("--out", "out_path", default=None, type=click.Path(path_type=Path),
              help="Write results to .csv or .xlsx. Omit for stdout.")
@click.option("--sheet-as-table", "sheet_as_table", is_flag=True, default=True,
              help="Register every sheet as a table (default).")
@common_output_options
def sql(src: Path, query: str, out_path: Path | None, sheet_as_table: bool,
        force: bool, backup: bool, as_json: bool, dry_run: bool,
        quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Run SQL over sheets (DuckDB if available, else SQLite)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    if dry_run:
        click.echo(f"would run SQL on {src}")
        return

    wb = load_workbook(src, read_only=True, data_only=True)

    sheets: dict[str, tuple[list[str], list[list]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_sanitize_col(h, i) for i, h in enumerate(rows[0])]
        data = [list(r) for r in rows[1:]]
        sheets[name] = (headers, data)

    try:
        import duckdb  # noqa: F401
        result_rows = _run_duckdb(query, sheets)
    except ImportError:
        result_rows = _run_sqlite(query, sheets)

    if out_path is None:
        if as_json:
            emit_json(result_rows)
        else:
            write_rows_csv("-", result_rows)
        return

    suffix = out_path.suffix.lower()
    if suffix == ".csv":
        text = _rows_to_csv_text(result_rows)

        def _writer(f):
            f.write(text.encode("utf-8"))

        safe_write(out_path, _writer, force=force, backup=backup, mkdir=mkdir)
    elif suffix in (".xlsx", ".xlsm"):
        _write_xlsx(out_path, result_rows, force=force, backup=backup, mkdir=mkdir)
    else:
        die(f"unsupported --out suffix: {suffix!r}", code=EXIT_INPUT, as_json=as_json)

    if as_json:
        emit_json({"path": str(out_path), "rows": len(result_rows)})
    elif not quiet:
        click.echo(f"wrote {out_path} ({len(result_rows)} rows)")


def _sanitize_col(val, i: int) -> str:
    if val is None or str(val).strip() == "":
        return f"col{i + 1}"
    return str(val)


def _run_duckdb(query: str, sheets):
    import duckdb

    con = duckdb.connect(":memory:")
    for name, (headers, data) in sheets.items():
        table = _quote_ident(name, duck=True)
        cols = ",".join(f"{_quote_ident(h, duck=True)} VARCHAR" for h in headers)
        con.execute(f"CREATE TABLE {table} ({cols})")
        if data:
            placeholders = "(" + ",".join(["?"] * len(headers)) + ")"
            rows = [tuple(list(r) + [None] * (len(headers) - len(r))) for r in data]
            con.executemany(f"INSERT INTO {table} VALUES {placeholders}", rows)
    cur = con.execute(query)
    cols = [d[0] for d in cur.description] if cur.description else []
    rows = cur.fetchall()
    return [dict(zip(cols, r)) for r in rows]


def _run_sqlite(query: str, sheets):
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    for name, (headers, data) in sheets.items():
        table = _quote_ident(name, duck=False)
        cols = ",".join(_quote_ident(h, duck=False) for h in headers)
        placeholders = ",".join(["?"] * len(headers))
        con.execute(f'CREATE TABLE {table} ({cols})')
        rows = [tuple(list(r) + [None] * (len(headers) - len(r))) for r in data]
        if rows:
            con.executemany(f"INSERT INTO {table} VALUES ({placeholders})", rows)
    try:
        cur = con.execute(query)
    except sqlite3.Error as e:
        die(f"SQL error: {e}", code=EXIT_INPUT)
    return [dict(row) for row in cur.fetchall()]


def _quote_ident(name: str, *, duck: bool) -> str:
    safe = name.replace('"', '""')
    return f'"{safe}"'


def _rows_to_csv_text(rows: list[dict]) -> str:
    import csv
    import io
    if not rows:
        return ""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue()


def _write_xlsx(path: Path, rows: list[dict], *, force: bool, backup: bool, mkdir: bool):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])

    def _save(f):
        wb.save(f)

    safe_write(path, _save, force=force, backup=backup, mkdir=mkdir)
