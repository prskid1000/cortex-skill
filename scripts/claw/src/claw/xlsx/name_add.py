"""claw xlsx name add — create a defined name (workbook- or sheet-scoped)."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import EXIT_INPUT, common_output_options, die, emit_json, safe_write


@click.group(name="name")
def name() -> None:
    """Defined-name operations."""


@name.command(name="add")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--name", "defined_name", required=True,
              help="The defined name (no spaces).")
@click.option("--refers-to", "refers_to", required=True,
              help='Formula or A1 reference, e.g. "=Sheet1!$B$2:$B$100".')
@click.option("--scope", default="workbook",
              help='"workbook" or a sheet name (for sheet-scoped names).')
@common_output_options
def name_add(src: Path, defined_name: str, refers_to: str, scope: str,
             force: bool, backup: bool, as_json: bool, dry_run: bool,
             quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Register a defined name pointing at --refers-to."""
    try:
        from openpyxl import load_workbook
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    if dry_run:
        click.echo(f"would add defined name {defined_name}={refers_to} (scope={scope})")
        return

    wb = load_workbook(src)
    value = refers_to if refers_to.startswith("=") else refers_to
    if value.startswith("="):
        value = value[1:]

    dn = DefinedName(name=defined_name, attr_text=value)

    if scope == "workbook":
        wb.defined_names[defined_name] = dn
    else:
        if scope not in wb.sheetnames:
            die(f"scope sheet not found: {scope!r}", code=EXIT_INPUT, as_json=as_json)
        sheet_idx = wb.sheetnames.index(scope)
        dn.localSheetId = sheet_idx
        wb[scope].defined_names[defined_name] = dn

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "name": defined_name,
                   "refers_to": refers_to, "scope": scope})
    elif not quiet:
        click.echo(f"added defined name {defined_name} -> {refers_to}")
