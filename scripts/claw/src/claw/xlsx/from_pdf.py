"""claw xlsx from-pdf — extract tables from a PDF into a multi-sheet workbook."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, PageSelector, common_output_options, die, emit_json, safe_write,
)


@click.command(name="from-pdf")
@click.argument("out", type=click.Path(path_type=Path))
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--pages", default="all",
              help='Page range (e.g. "1-5", "all", "odd", "2-end").')
@click.option("--strategy", default="lines",
              type=click.Choice(["lines", "text"]),
              help="pdfplumber table-extraction strategy.")
@click.option("--sheet-per-page", "sheet_per_page", is_flag=True,
              help="Merge all tables on a page into one sheet.")
@click.option("--snap-tol", "snap_tol", default=3, type=int)
@common_output_options
def from_pdf(out: Path, src: Path, pages: str, strategy: str,
             sheet_per_page: bool, snap_tol: int,
             force: bool, backup: bool, as_json: bool, dry_run: bool,
             quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Convert PDF tables to sheets via pdfplumber."""
    try:
        import pdfplumber
        from openpyxl import Workbook
    except ImportError:
        die("pdfplumber and openpyxl required", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx,pdf]'", as_json=as_json)

    if dry_run:
        click.echo(f"would write {out} from {src} pages={pages}")
        return

    table_settings = {
        "vertical_strategy": strategy,
        "horizontal_strategy": strategy,
        "snap_tolerance": snap_tol,
    }

    wb = Workbook()
    wb.remove(wb.active)

    sheets: list[str] = []
    seen: set[str] = set()

    with pdfplumber.open(str(src)) as pdf:
        total = len(pdf.pages)
        page_nums = PageSelector(pages).resolve(total)
        for pn in page_nums:
            page = pdf.pages[pn - 1]
            try:
                extracted = page.extract_tables(table_settings) or []
            except Exception:
                extracted = []
            if not extracted:
                continue

            if sheet_per_page:
                name = f"Page{pn}"
                name = _unique(name, seen)
                ws = wb.create_sheet(title=name)
                for ti, tbl in enumerate(extracted):
                    if ti > 0:
                        ws.append([])
                    for row in tbl:
                        ws.append([(c if c is not None else "") for c in row])
                sheets.append(name)
            else:
                for ti, tbl in enumerate(extracted, start=1):
                    name = f"Page{pn}_T{ti}" if len(extracted) > 1 else f"Page{pn}"
                    name = _unique(name, seen)
                    ws = wb.create_sheet(title=name)
                    for row in tbl:
                        ws.append([(c if c is not None else "") for c in row])
                    sheets.append(name)

    if not sheets:
        die("no tables extracted", code=EXIT_INPUT, as_json=as_json)

    def _save(f):
        wb.save(f)

    safe_write(out, _save, force=force, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(out), "sheets": sheets})
    elif not quiet:
        click.echo(f"wrote {out} with {len(sheets)} sheet(s)")


def _unique(name: str, seen: set[str]) -> str:
    name = name[:31]
    base = name
    n = 2
    while name in seen:
        suffix = f"({n})"
        name = (base[:31 - len(suffix)] + suffix)
        n += 1
    seen.add(name)
    return name
