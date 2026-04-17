"""claw xlsx pivots list — enumerate pivot tables in a workbook (read-only)."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import EXIT_INPUT, die, emit_json


@click.group(name="pivots")
def pivots() -> None:
    """Pivot-table operations (read-only — openpyxl cannot author pivots)."""


@pivots.command(name="list")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--json", "as_json", is_flag=True)
def pivots_list(src: Path, as_json: bool) -> None:
    """List pivot tables present in the workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    wb = load_workbook(src)
    results: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        pivots_attr = getattr(ws, "_pivots", None)
        if not pivots_attr:
            continue
        for p in pivots_attr:
            cache = getattr(p, "cache", None)
            cache_ref = None
            if cache is not None:
                cache_src = getattr(cache, "cacheSource", None)
                if cache_src is not None:
                    wsrc = getattr(cache_src, "worksheetSource", None)
                    if wsrc is not None:
                        cache_ref = f"{wsrc.sheet}!{wsrc.ref}" if wsrc.sheet else wsrc.ref
            results.append({
                "sheet": sheet_name,
                "name": getattr(p, "name", None),
                "location": getattr(getattr(p, "location", None), "ref", None),
                "data_source": cache_ref,
            })

    if as_json:
        emit_json(results)
    else:
        if not results:
            click.echo("(no pivot tables)")
            return
        for r in results:
            click.echo(f"{r['sheet']}: {r['name']} @ {r['location']} ← {r['data_source']}")
