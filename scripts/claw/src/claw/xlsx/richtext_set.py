"""claw xlsx richtext set — write a CellRichText run list to a cell."""

from __future__ import annotations

import json
from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, common_output_options, die, emit_json, read_text, safe_write,
)


@click.group(name="richtext")
def richtext() -> None:
    """Rich-text cell operations."""


@richtext.command(name="set")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True)
@click.option("--cell", required=True, help="Target cell in A1 notation.")
@click.option("--runs", "runs_src", required=True,
              help="Path to JSON, '-' for stdin, or inline JSON string.")
@common_output_options
def richtext_set(src: Path, sheet: str, cell: str, runs_src: str,
                 force: bool, backup: bool, as_json: bool, dry_run: bool,
                 quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Replace a cell's contents with a CellRichText list of formatted runs."""
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        die("openpyxl not installed (need >=3.1 for CellRichText)",
            code=EXIT_INPUT, hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    runs_raw: str
    if runs_src.strip().startswith("["):
        runs_raw = runs_src
    else:
        runs_raw = read_text(runs_src)

    try:
        runs = json.loads(runs_raw)
    except json.JSONDecodeError as e:
        die(f"invalid --runs JSON: {e}", code=EXIT_INPUT, as_json=as_json)

    if not isinstance(runs, list) or not runs:
        die("--runs must be a non-empty JSON array", code=EXIT_INPUT, as_json=as_json)

    if dry_run:
        click.echo(f"would set {sheet}!{cell} with {len(runs)} run(s)")
        return

    wb = load_workbook(src)
    ws = wb[sheet]

    crt = CellRichText()
    for run in runs:
        if not isinstance(run, dict) or "text" not in run:
            die("each run needs a 'text' field", code=EXIT_INPUT, as_json=as_json)
        font_kwargs = {}
        if run.get("bold"):
            font_kwargs["b"] = True
        if run.get("italic"):
            font_kwargs["i"] = True
        if "color" in run and run["color"]:
            font_kwargs["color"] = run["color"].lstrip("#")
        if "size" in run and run["size"]:
            font_kwargs["sz"] = run["size"]
        if "font" in run and run["font"]:
            font_kwargs["rFont"] = run["font"]
        if font_kwargs:
            crt.append(TextBlock(InlineFont(**font_kwargs), run["text"]))
        else:
            crt.append(run["text"])

    ws[cell] = crt

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "sheet": sheet, "cell": cell,
                   "runs": len(runs)})
    elif not quiet:
        click.echo(f"set rich text on {sheet}!{cell} ({len(runs)} run(s))")
